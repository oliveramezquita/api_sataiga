from collections import Counter
from typing import Any, Dict, List, Tuple, Set, Optional
from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException
from rest_framework.exceptions import ValidationError
from api.serializers.file_serializer import FileUploadSerializer
from api.helpers.formats import norm
from api.constants import ALLOWED_LAID
from api.services.base_service import BaseService
from api.repositories.lot_repository import LotRepository
from api.repositories.home_production_repository import HomeProductionRepository
from api.repositories.prototype_repository import PrototypeRepository
from api.serializers.lot_serializer import LotSerializer
from api.functions.explosion import explosion


class LotService(BaseService):
    CACHE_PREFIX = "lots"

    def __init__(self):
        self.lot_repo = LotRepository()
        self.hp_repo = HomeProductionRepository()
        self.proto_repo = PrototypeRepository()

    @staticmethod
    def _default_progress() -> Dict[str, Any]:
        def block():
            return {"stages": [], "percentage": 0.0}

        return {
            "cocina": block(),
            "closet": block(),
            "puertas": block(),
            "mdeb": block(),
            "waldras": block(),
            "instalacion": block(),
        }

    @staticmethod
    def split_and_process_lots(data: Any) -> Tuple[List[Dict[str, Any]], List[Dict[str, Any]]]:
        insertions: List[Dict[str, Any]] = []
        updates: List[Dict[str, Any]] = []

        if data is None:
            return insertions, updates
        if not isinstance(data, list):
            raise ValidationError("lots debe ser una lista.")

        for item in data:
            if not isinstance(item, dict):
                continue  # o ValidationError

            _id = item.get("_id")
            if _id:
                updates.append(item)
            else:
                insertions.append(item)

        return insertions, updates

    def _process_workbook(
        self,
        hp_id: str,
        wb,
        allowed_prototypes: Set[str],
    ) -> Tuple[List[Dict[str, str]], List[Dict[str, Any]]]:
        ws = wb.active
        valid_rows: List[Dict[str, str]] = []
        errors: List[Dict[str, Any]] = []

        for row_idx in range(2, ws.max_row + 1):
            block = ws.cell(row=row_idx, column=1).value
            lot = ws.cell(row=row_idx, column=2).value
            laid = ws.cell(row=row_idx, column=3).value
            prototype = ws.cell(row=row_idx, column=4).value

            # Ignorar filas totalmente vacías
            if block is None and lot is None and laid is None and prototype is None:
                continue

            # Todos obligatorios
            missing = []
            if block is None:
                missing.append("manzana")
            if lot is None:
                missing.append("lote")
            if laid is None:
                missing.append("sembrado")
            if prototype is None:
                missing.append("prototipo")

            if missing:
                errors.append(
                    {"row": row_idx,
                        "error": f"Campos requeridos faltantes: {', '.join(missing)}"}
                )
                continue

            block_s = str(block).strip()
            lot_s = str(lot).strip()
            laid_s = str(laid).strip().upper()
            prototype_s = str(prototype).strip()

            empty = []
            if not block_s:
                empty.append("manzana")
            if not lot_s:
                empty.append("lote")
            if not laid_s:
                empty.append("sembrado")
            if not prototype_s:
                empty.append("prototipo")

            if empty:
                errors.append(
                    {"row": row_idx,
                        "error": f"Campos vacíos: {', '.join(empty)}"}
                )
                continue

            if laid_s not in ALLOWED_LAID:
                errors.append(
                    {
                        "row": row_idx,
                        "error": f"Sembrado inválido: '{laid_s}'. Solo se permite: DERECHO o IZQUIERDO",
                    }
                )
                continue

            if norm(prototype_s) not in allowed_prototypes:
                errors.append(
                    {
                        "row": row_idx,
                        "error": f"El prototipo no existe para esta OD: '{prototype_s}'",
                    }
                )
                continue

            existing_lot = self.lot_repo.find_one({
                'home_production_id': hp_id,
                'block': block_s,
                'lot': lot_s,
                'laid': laid_s,
                'prototype': prototype_s,
            })
            if not existing_lot:
                valid_rows.append(
                    {"block": block_s, "lot": lot_s,
                        "laid": laid_s, "prototype": prototype_s}
                )

        return valid_rows, errors

    def _update_hp_lots(self, home_production_id: str, lots: List[Dict[str, Any]]):
        if home_production_id and lots:
            prototype_counter = Counter()
            for lot in lots:
                prototype_counter[lot['prototype']] += 1

            result = {
                'total': len(lots),
                'prototypes': dict(prototype_counter),
            }

            total_sum = sum(float(l["percentage"]) for l in lots)
            self._update(
                repo=self.hp_repo,
                _id=home_production_id,
                data={
                    'lots': result,
                    'progress': round(total_sum / len(lots), 2)},
                cache_prefix='home_production'
            )

    def create(self, home_production_id: str, data: Dict[str, Any]):
        if not home_production_id:
            raise ValidationError("home_production_id es requerido.")
        hp = self.hp_repo.find_by_id(home_production_id) or {}
        prev_lots = hp.get('lots', {})
        lots = data.get("lots")
        if not lots:
            # respuesta consistente
            return {"success": [], "errors": []}

        insertions, updates = self.split_and_process_lots(lots)

        errors: List[Dict[str, Any]] = []

        # -------- Updates --------
        for lot in updates:
            _id = lot.get("_id")
            if not _id:
                errors.append(
                    {"lot": lot, "error": "Falta _id para actualizar."})
                continue

            # no mutar input
            payload = {k: v for k, v in lot.items() if k != "_id"}

            # fuerza relación (evita que alguien cambie HP)
            payload["home_production_id"] = home_production_id

            try:
                self._update(
                    repo=self.lot_repo,
                    _id=_id,
                    data=payload,
                    cache_prefix=self.CACHE_PREFIX,
                )
            except Exception as e:
                errors.append(lot)

        for lot in insertions:
            payload = {
                "home_production_id": home_production_id,
                **lot,
                "percentage": 0,
                "progress": self._default_progress(),
            }

            try:
                self._create(
                    repo=self.lot_repo,
                    data=payload,
                    required_fields=["prototype", "block", "lot", "laid"],
                    cache_prefix=self.CACHE_PREFIX,
                )
            except Exception as e:
                errors.append(lot)

        updated_lots = self.lot_repo.find_all(
            {"home_production_id": home_production_id})
        self._update_hp_lots(home_production_id, updated_lots)
        explosion.delay(home_production_id, int(prev_lots.get('total', 0)))
        return {
            "success": LotSerializer(updated_lots, many=True).data,
            "errors": errors,
        }

    def upload(self, home_production_id: str, data, request_file):
        if not home_production_id:
            raise ValidationError("home_production_id es requerido.")

        hp = self.hp_repo.find_by_id(home_production_id)
        if not hp:
            raise LookupError("La OD no existe.")
        lots = hp.get('lots', {})
        protos = self.proto_repo.find_all(
            {"client_id": hp.get("client_id"), "front": hp.get("front")}
        ) or []
        allowed_prototypes: Set[str] = {
            norm(p.get("name")) for p in protos if isinstance(p, dict) and p.get("name")
        }

        serializer = FileUploadSerializer(data=data)
        if not serializer.is_valid():
            raise ValidationError(serializer.errors)

        if not request_file:
            raise ValidationError("El archivo es requerido.")

        try:
            if hasattr(request_file, "seek"):
                request_file.seek(0)
            workbook = load_workbook(request_file, data_only=True)
        except (InvalidFileException, OSError, ValueError) as e:
            raise ValidationError(f"Archivo Excel inválido: {e}")

        valid_rows, errors = self._process_workbook(
            home_production_id,
            workbook,
            allowed_prototypes)

        if not valid_rows and errors:
            raise ValidationError(
                {"detail": "No hay filas válidas.", "errors": errors})

        for row in valid_rows:
            self._create(
                repo=self.lot_repo,
                data={
                    "home_production_id": home_production_id,
                    **row,
                    "percentage": 0,
                    "progress": self._default_progress(),
                },
                required_fields=["prototype", "block", "lot", "laid"],
                cache_prefix=self.CACHE_PREFIX,
            )

        updated_lots = self.lot_repo.find_all(
            {"home_production_id": home_production_id})
        self._update_hp_lots(home_production_id, updated_lots)
        explosion.delay(home_production_id, int(lots.get('total', 0)))
        return {
            "success": LotSerializer(updated_lots, many=True).data,
            "errors": errors,
        }

    def update(self, home_production_id: str, data: Dict[str, Any]) -> str:
        if not home_production_id:
            raise ValidationError("home_production_id es requerido.")

        if not data:
            raise ValidationError(
                "Se requieren datos para actualizar los lotes.")

        self.lot_repo.upsert_one(
            {'home_production_id': home_production_id}, data)
        updated_lots = self.lot_repo.find_all(
            {"home_production_id": home_production_id})
        self._update_hp_lots(home_production_id, updated_lots)

        return 'Lote actualizado correctamente.'

    def delete(self, lot_id: str) -> str:
        if not lot_id:
            raise ValidationError("lot_id es requerido.")
        lot = self._get_by_id(repo=self.lot_repo, _id=lot_id)
        if not lot:
            raise LookupError("La lote no existe.")

        self._delete(
            repo=self.lot_repo,
            _id=lot_id,
            cache_prefix=self.CACHE_PREFIX
        )

        home_production_id = lot.get('home_production_id')
        updated_lots = self.lot_repo.find_all(
            {"home_production_id": home_production_id})
        self._update_hp_lots(home_production_id, updated_lots)
        explosion.delay(home_production_id, len(updated_lots) + 1)

        return 'Lote eliminado correctamente.'

    def get(self, home_production_id: Optional[str] = None) -> List[Dict[str, Any]]:
        filters = {
            "home_production_id": home_production_id} if home_production_id else None
        lots = self._get_all_cached(
            repo=self.lot_repo,
            filters=filters,
            prefix=self.CACHE_PREFIX
        ) or []
        return LotSerializer(lots, many=True).data
