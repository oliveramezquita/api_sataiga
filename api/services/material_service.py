from io import BytesIO
from pathlib import Path, PurePath
from typing import List, Tuple
from django.conf import settings
from openpyxl import load_workbook
from api.services.base_service import BaseService
from api.repositories.material_repository import MaterialRepository
from api.serializers.material_serializer import MaterialSerializer


class MaterialService(BaseService):
    CACHE_PREFIX = "materials"
    DEFAULT_TEMPLATE_PATH = (
        Path(settings.BASE_DIR) / "media" / "materials" / "templates"
    )

    def __init__(self):
        self.material_repo = MaterialRepository()

    def get(self, filters: dict, sort_by: str = 'concept', order_by: int = 1):
        items = self._get_all_cached(
            self.material_repo,
            filters,
            prefix=self.CACHE_PREFIX,
            order_field=sort_by,
            order=order_by
        )

        return MaterialSerializer(items, many=True).data

    def get_paginated(self, filters: dict, page: int, page_size: int, sort_by: str = 'concept', order_by: int = 1):
        items = self._get_all_cached(
            self.material_repo,
            filters,
            prefix=self.CACHE_PREFIX,
            order_field=sort_by,
            order=order_by
        )
        return self._paginate(items, page, page_size, serializer=MaterialSerializer)

    def export_format(
        self,
        filename: str,
        material_ids: List[str] | None = None,
    ) -> Tuple[bytes, str]:
        if material_ids and not all(isinstance(x, str) for x in material_ids):
            raise TypeError("material_ids debe ser una lista de strings")

        # Seguridad: solo permitir nombres de archivo, no rutas
        if PurePath(filename).name != filename:
            raise ValueError("Nombre de archivo inválido.")

        template_path = self.DEFAULT_TEMPLATE_PATH / filename

        if not template_path.exists():
            raise FileNotFoundError(f"No existe la plantilla: {template_path}")

        wb = load_workbook(filename=str(template_path))
        ws = wb.active
        start_row = 3

        if material_ids:
            materials = self.material_repo.find_many_by_ids(material_ids)

            by_id = {str(m.get("_id")): m for m in materials}
            ordered = [by_id[mid] for mid in material_ids if mid in by_id]

            for i, m in enumerate(ordered):
                row = start_row + i
                ws[f"A{row}"] = m.get("concept") or ""
                ws[f"B{row}"] = m.get("measurement") or ""
                ws[f"C{row}"] = m.get("sku") or ""

        bio = BytesIO()
        wb.save(bio)
        bio.seek(0)
        return bio.read(), filename
