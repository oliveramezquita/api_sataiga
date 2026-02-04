import uuid
from typing import Any, Dict, List, Tuple, Mapping, Optional
from openpyxl import load_workbook
from rest_framework.exceptions import ValidationError
from openpyxl.utils.exceptions import InvalidFileException
from api.helpers.formats import to_float, mongo_to_json, to_number, normalize_num
from api.services.base_service import BaseService
from api.repositories.client_repository import ClientRepository
from api.repositories.volumetry_repository import VolumetryRepository
from api.repositories.material_repository import MaterialRepository
from api.serializers.file_serializer import FileUploadSerializer


class VolumetryService(BaseService):
    """Lógica de negocio para la volumetría."""

    CACHE_PREFIX = "volumetries"

    def __init__(self):
        self.client_repo = ClientRepository()
        self.volumetry_repo = VolumetryRepository()
        self.material_repo = MaterialRepository()

    @staticmethod
    def validate_and_sum_volumetry(rows: Any) -> Tuple[List[Dict[str, Any]], float]:
        """
        - Convierte factory/installation/delivery a float (>=0)
        - Recalcula total_x = factory + installation + delivery
        - Retorna (rows_limpias, suma_total_x)
        """
        if rows is None:
            rows = []
        if not isinstance(rows, list):
            raise ValueError("volumetry debe ser una lista.")

        cleaned_rows: List[Dict[str, Any]] = []
        total_sum = 0.0

        for idx, row in enumerate(rows):
            if not isinstance(row, Mapping):
                raise ValueError(f"volumetry[{idx}] debe ser un objeto.")

            # Campos permitidos (whitelist)
            row_id = row.get("id")
            area = row.get("area")

            # Si quieres exigir area:
            if area is None or (isinstance(area, str) and not area.strip()):
                raise ValueError(f"volumetry[{idx}].area no puede ser vacío.")

            # Usando tu to_float mejorado con min_value keyword-only
            factory = to_float(row.get("factory"), 0.0, min_value=0.0)
            installation = to_float(
                row.get("installation"), 0.0, min_value=0.0)
            delivery = to_float(row.get("delivery"), 0.0, min_value=0.0)

            total_x = factory + installation + delivery
            total_sum += total_x

            cleaned_rows.append({
                "id": row_id,
                "area": area.strip() if isinstance(area, str) else area,
                "factory": factory,
                "installation": installation,
                "delivery": delivery,
                "total_x": total_x,
            })

        return cleaned_rows, total_sum

    @staticmethod
    def process_workbook(wb):
        ws = wb.active

        areas = []
        col = 4  # D
        max_col = ws.max_column

        while col <= max_col:
            name = ws.cell(row=1, column=col).value
            if name is None:
                col += 1
                continue

            name = str(name).strip()
            if name.upper() == "COCINA":
                areas.append(
                    ("COCINA", {"factory": col, "installation": col + 1, "delivery": col + 2}))
                col += 3
            else:
                areas.append((name, {"factory": col, "installation": col + 1}))
                col += 2

        # 2) Procesar SKUs (col C) desde fila 3
        result = {}
        for row in range(3, ws.max_row + 1):
            sku = ws.cell(row=row, column=3).value  # C

            # si la fila está vacía, la saltamos
            if sku is None or str(sku).strip() == "":
                continue

            sku = str(sku).strip()
            per_area = []

            for area_name, cols in areas:
                item = {
                    "area": area_name,
                    "factory": normalize_num(to_number(ws.cell(row=row, column=cols["factory"]).value)),
                    "installation": normalize_num(to_number(ws.cell(row=row, column=cols["installation"]).value)),
                }
                if "delivery" in cols:
                    item["delivery"] = normalize_num(
                        to_number(ws.cell(row=row, column=cols["delivery"]).value))

                per_area.append(item)

            result[sku] = per_area

        return result

    def _process_data(self, client_id: str, front: str, prototype: str, data: Dict[str, List[Dict]]):
        num_inserted = 0
        num_updated = 0
        errors = []
        processed_volumetry = []

        for sku, v in (data or {}).items():
            material = self.material_repo.find_one({"sku": sku})

            if not material:
                # Yo no lo silencaría: lo reporto
                errors.append({
                    "sku": sku,
                    "message": f"No existe material con SKU: {sku}"
                })
                continue

            material_id = str(material.get("_id"))
            if not material_id:
                errors.append({
                    "sku": sku,
                    "message": f"Material encontrado pero sin _id para SKU: {sku}"
                })
                continue

            volumetry = []
            for item in v:
                volumetry.append({
                    "id": str(uuid.uuid4()),
                    "area": item.get("area"),
                    "factory": item.get("factory", 0),
                    "installation": item.get("installation", 0),
                    "delivery": item.get("delivery", 0),
                })

            try:
                rows, total = self.validate_and_sum_volumetry(volumetry)
            except ValueError as e:
                # Mejor como ValidationError, pero si no quieres tocar validate_and_sum_volumetry:
                errors.append({"sku": sku, "message": str(e)})
                continue

            query = {
                "client_id": client_id,
                "front": front,
                "prototype": prototype,
                "material_id": material_id,
            }

            set_data = {
                **query,
                "supplier_id": material.get("supplier_id"),
                "volumetry": rows,
                "total": round(total, 2),
            }

            # Ideal: que upsert_one te regrese inserted/updated
            doc = self.volumetry_repo.upsert_one(query, set_data)
            processed_volumetry.append({**doc,
                                        'id': material_id,
                                        'concept': material.get('concept'),
                                        'sku': material.get('sku'),
                                        'division': material.get('division'),
                                        'measurement': material.get('measurement'),
                                        'presentation': material.get('presentation')})
            created = doc.get("created_at")
            updated = doc.get("updated_at")

            if created and updated and created == updated:
                num_inserted += 1
            else:
                num_updated += 1

        return {
            "num_inserted": num_inserted,
            "num_updated": num_updated,
            "errors": errors,
            "volumetry": [mongo_to_json(v) for v in processed_volumetry]}

    def load_volumetry(self, payload: Dict[str, Any]):
        """
        Carga la volumetría asociada a un cliente/front/prototype.
        Para cada material hace upsert por (client_id, front, prototype, material_id).
        """
        if not isinstance(payload, dict):
            raise ValueError("payload inválido.")

        client_id = payload.get("client_id")
        front = payload.get("front")
        prototype = payload.get("prototype")
        items = payload.get("volumetry")

        # Validaciones mínimas
        if not client_id:
            raise ValueError("client_id es requerido.")
        if not front:
            raise ValueError("front es requerido.")
        if not prototype:
            raise ValueError("prototype es requerido.")
        if not isinstance(items, list):
            raise ValueError("volumetry debe ser una lista.")

        client = self.client_repo.find_by_id(client_id)
        if not client:
            raise LookupError("El cliente no existe.")

        for idx, item in enumerate(items):
            if not isinstance(item, Mapping):
                raise ValueError(f"volumetry[{idx}] debe ser un objeto.")

            material_id = item.get("material_id")
            supplier_id = item.get("supplier_id")

            if not material_id:
                raise ValueError(f"volumetry[{idx}].material_id es requerido.")
            if not supplier_id:
                raise ValueError(f"volumetry[{idx}].supplier_id es requerido.")

            material = self.material_repo.find_by_id(material_id)
            if not material:
                raise LookupError(f"El material no existe: {material_id}")

            rows, total = self.validate_and_sum_volumetry(
                item.get("volumetry"))

            # Clave natural (única) SIN supplier_id
            query = {
                "client_id": client_id,
                "front": front,
                "prototype": prototype,
                "material_id": material_id,
            }

            set_data = {
                **query,
                "supplier_id": supplier_id,
                "volumetry": rows,
                "total": total,
            }

            # Upsert atómico (reduce race conditions)
            self.volumetry_repo.upsert_one(query, set_data)

        # Si tienes caché por prefijo y quieres invalidar todo lo de ese cliente/front/prototype,
        # aquí podrías limpiar caché. Depende de cómo lo maneje BaseService.
        return True

    def upload(self, client_id: str, front: str, prototype: str, data, request_file):
        if not client_id:
            raise ValidationError("client_id es requerido.")
        if not front:
            raise ValidationError("front es requerido.")
        if not prototype:
            raise ValidationError("prototype es requerido.")
        client = self.client_repo.find_by_id(client_id)
        if not client:
            raise LookupError("El cliente no existe.")

        serializer = FileUploadSerializer(data=data)
        if not serializer.is_valid():
            raise ValidationError(serializer.errors)

        if not request_file:
            raise ValidationError("El archivo es requerido.")

        try:
            workbook = load_workbook(request_file, data_only=True)
        except (InvalidFileException, OSError, ValueError) as e:
            raise ValidationError(f"Archivo Excel inválido: {e}")

        volumetry_data = self.process_workbook(workbook)

        # Si quieres validar que venga al menos un SKU
        if not volumetry_data:
            raise ValidationError(
                "El archivo no contiene SKUs en la columna C a partir de la fila 3.")

        summary = self._process_data(
            client_id, front, prototype, volumetry_data)

        return summary

    def get(
            self,
            client_id: Optional[str] = None,
            front: Optional[str] = None,
            prototype: Optional[str] = None):
        """
        Obtiene la volumetría aplicando filtros opcionales y
        agrega datos del material (concept, sku, measurement, presentation)
        al nivel raíz para evitar N+1 queries en serializer.
        """
        filters = {k: v for k, v in {
            "client_id": client_id,
            "front": front,
            "prototype": prototype,
        }.items() if v is not None}

        volumetries = self._get_all_cached(
            repo=self.volumetry_repo,
            filters=filters,
            prefix=self.CACHE_PREFIX,
        ) or []

        # 1) Recolectar material_ids (strings)
        material_ids = [v.get("material_id")
                        for v in volumetries if v.get("material_id")]

        # 2) Prefetch en 1 query (usando el repo)
        mats = self.material_repo.find_many_by_ids(
            material_ids,
            projection={"concept": 1, "sku": 1, "division": 1,
                        "measurement": 1, "presentation": 1},
            dedupe=True,
        ) or []

        materials_map: Dict[str, Dict[str, Any]] = {
            str(m["_id"]): m for m in mats}

        # 3) Aplanar campos del material en cada volumetría
        for v in volumetries:
            m = materials_map.get(v.get("material_id"), {})
            v['id'] = v['material_id']
            v["concept"] = m.get("concept")
            v["sku"] = m.get("sku")
            v["division"] = m.get("division")
            v["measurement"] = m.get("measurement")
            v["presentation"] = m.get("presentation")

        return [mongo_to_json(v) for v in volumetries]
