import traceback
import pytz
import json
from urllib.parse import parse_qs
from api.constants import DEFAULT_PAGE_SIZE
from api_sataiga.handlers.mongodb_handler import MongoDBHandler
from django.core.paginator import Paginator
from api.helpers.http_responses import ok_paginated, ok, created, bad_request, not_found
from api.serializers.inbound_serializer import InboundSerializer
from api.serializers.inventory_quantity_serializer import InventoryQuantitySerializer
from bson import ObjectId
from api.helpers.validations import objectid_validation
from datetime import datetime, timedelta
from api.serializers.file_serializer import FileUploadSerializer
from openpyxl import load_workbook
from decimal import Decimal, InvalidOperation


class InboundUseCase:
    def __init__(self, request=None, **kwargs):
        if request:
            params = parse_qs(request.META['QUERY_STRING'])
            self.page = params['page'][0] if 'page' in params else 1
            self.page_size = params['itemsPerPage'][0] \
                if 'itemsPerPage' in params else DEFAULT_PAGE_SIZE
            self.q = params['q'][0] if 'q' in params else None
            self.supplier = params['supplier'][0] if 'supplier' in params else None
            self.created_at = params['created_at'][0] if 'created_at' in params else None
        self.data = kwargs.get('data', None)
        self.id = kwargs.get('id', None)
        self.project_type = kwargs.get('project_type', None)
        self.material_id = kwargs.get('material', None)

    def __perform_counting(self, db, project, items, inbound_id):
        for item in items:
            material = {
                'id': item['material_id'],
                'color': item.get('color', ''),
                'concept': item['concept'],
                'measurement': item['measurement'],
                'sku': item['sku'],
                'division': item['division'],
                'supplier_id': item['supplier_id'],
                'supplier_code': item.get('supplier_code', ''),
                'inventory_price': item.get('inventory_price', ''),
                'market_price': item.get('market_price', ''),
                'presentation': item.get('presentation', ''),
                'reference': item.get('reference', ''),
            }

            inventory = MongoDBHandler.find(
                db, 'inventory', {'material.id': item['material_id']})
            inventory_id = None
            if inventory and len(inventory) > 0:
                inventory_id = str(inventory[0]['_id'])
                quantity = float(inventory[0]['quantity']) + \
                    float(item['delivered']['quantity'])
                MongoDBHandler.modify(db, 'inventory', {'_id': inventory[0]['_id']}, {
                    'quantity': round(float(quantity), 2)
                })
            if not inventory_id:
                inventory_id = MongoDBHandler.record(db, 'inventory', {
                    'material': material,
                    'quantity': item['delivered']['quantity'],
                })
            MongoDBHandler.record(db, 'inventory_quantity', {
                'inventory_id': str(inventory_id),
                'inbound_id': str(inbound_id),
                'material_id': material['id'],
                'project': project,
                'quantity': round(float(item['delivered']['quantity']), 2),
                'rack': item['delivered']['rack'],
                'level': item['delivered']['level'],
                'module': item['delivered']['module'],
                'status': 0,
            })

    def __get_home_production(self):
        results = []
        with MongoDBHandler('home_production') as db:
            home_production = db.extract()
            if home_production:
                results = [{"_id": str(
                    item["_id"]), "name": f"{item['front']} - {item['od']}"} for item in home_production]
        return results

    def __valid_value(self, data, idx):
        if len(data) <= idx:
            return None

        not_valid = ['', 'None']
        if not data[idx] or data[idx] in not_valid:
            return None

        if not isinstance(data[idx], str):
            return str(data[idx]).strip()

        return data[idx].strip()

    def __get_material(self, **kwargs):
        with MongoDBHandler('materials') as db:
            and_conditions = []

            supplier_id = kwargs.get('supplier_id')
            if not supplier_id:
                return False
            and_conditions.append({'supplier_id': supplier_id})

            supplier_code = kwargs.get('supplier_code')
            if supplier_code:
                and_conditions.append({'supplier_code': supplier_code})

            or_conditions = []
            sku = kwargs.get('sku')
            if sku:
                or_conditions.append(
                    {'sku': {'$regex': f'^{sku}', '$options': 'i'}})

            concept = kwargs.get('concept')
            if concept:
                or_conditions.append(
                    {'concept': {'$regex': f'^{concept}', '$options': 'i'}})

            if or_conditions:
                and_conditions.append({'$or': or_conditions})

            filters = {'$and': and_conditions} if len(
                and_conditions) > 1 else and_conditions[0]

            projection = {
                "color": 1,
                "concept": 1,
                "measurement": 1,
                "supplier_id": 1,
                "supplier_code": 1,
                "inventory_price": 1,
                "market_price": 1,
                "sku": 1,
                "presentation": 1,
                "reference": 1,
                "division": 1
            }

            material = db.extract(filters, projection=projection)

            if material:
                doc = material[0]
                doc["material_id"] = str(doc.pop("_id"))
                return doc
            return False

    def __process_workbook(self, workbook, supplier_id):
        sheet = workbook.active

        data = []
        errors = []

        for idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            row_errors = []

            def get_decimal(value, field_name):
                try:
                    if value is None or value == '':
                        return None
                    convert = Decimal(str(value)).quantize(Decimal('0.01'))
                    return str(convert)
                except (InvalidOperation, ValueError):
                    row_errors.append(
                        f"Fila {idx}: '{field_name}' no es un número decimal válido.")
                    return None

            sku = row[3]
            concept = row[4]
            supplier_code = self.__valid_value(row, 5)

            if not sku and not concept:
                row_errors.append(
                    f"Fila {idx}: 'SKU' o 'CONCEPTO' son obligatorios.")

            record = self.__get_material(
                supplier_id=supplier_id,
                sku=sku,
                concept=concept,
                supplier_code=supplier_code
            )
            if not record:
                row_errors.append(
                    f"No existe un material registrado que coincida con el concepto: {concept}, SKU: {sku} o código de proveedor: {supplier_code}.")
            else:
                record['delivered'] = {
                    'rack': self.__valid_value(row, 0),
                    'level': self.__valid_value(row, 1),
                    'module': self.__valid_value(row, 2),
                    'quantity': get_decimal(row[6], "CANTIDAD"),
                    'registration_date': datetime.now(pytz.timezone('America/Mexico_City')),
                }
                record['total_quantity'] = 0

            if row_errors:
                errors.append({"fila": idx, "errores": row_errors})
            else:
                data.append(record)

        return data, errors

    def __normalize_field(self, value):
        if value == "null":
            return None

        if isinstance(value, str):
            try:
                return json.loads(value)
            except json.JSONDecodeError:
                return value

        return value

    def __expand_items(self, data):
        expanded = []
        for record in data:
            base = {k: v for k, v in record.items() if k != "items"}

            for item in record.get("items", []):
                new_record = base.copy()
                new_record["item"] = item
                expanded.append(new_record)
        return expanded

    def get(self):
        with MongoDBHandler('inbounds') as db:
            # TODO - Add filters
            inbounds = db.extract()
            paginator = Paginator(inbounds, per_page=self.page_size)
            page = paginator.get_page(self.page)
            return ok_paginated(
                paginator,
                page,
                InboundSerializer(page.object_list, many=True).data
            )

    def get_project(self):
        results = []
        if self.project_type == 'Vivienda en Serie':
            results = self.__get_home_production()
        # TODO - Get from special projects
        return ok(results)

    def save(self):
        with MongoDBHandler('inbounds') as db:
            required_fields = ['project', 'items']
            if all(i in self.data for i in required_fields):
                # TODO - Calculate the total of new materials added
                data = self.data
                data['folio'] = db.set_next_folio('inbound')
                data['status'] = 0
                id = db.insert(data)
                return created({'id': str(id)})
            return bad_request('Algunos campos requeridos no han sido completados.')

    def update(self):
        with MongoDBHandler('inbounds') as db:
            inbound = db.extract(
                {'_id': ObjectId(self.id)}) if objectid_validation(self.id) else None
            if inbound:
                if 'items' in self.data:
                    self.__perform_counting(
                        db, inbound[0]['project'], self.data['items'], inbound[0]['_id'])
                db.update({'_id': ObjectId(self.id)}, self.data)
                return ok('La entrada de materiales fue almacenda correctamente.')
            return not_found('La entrada no existe o no hay materiales de entrada para almacenar.')

    def get_by_id(self):
        with MongoDBHandler('inbounds') as db:
            inbound = db.extract(
                {'_id': ObjectId(self.id)}) if objectid_validation(self.id) else None
            if inbound:
                return ok(InboundSerializer(inbound[0]).data)
            return not_found('La entrada no existe.')

    def get_by_material(self):
        query = {'status': {'$in': [0, 1, 2]}}
        created_at_param = None
        if self.created_at:
            created_at_param = self.created_at.replace(' to ', '+to+')

        if isinstance(created_at_param, str):
            try:
                if '+to+' in created_at_param:
                    start_str, end_str = created_at_param.split('+to+')
                    start_date = datetime.strptime(start_str, '%Y-%m-%d')
                    end_date = datetime.strptime(
                        end_str, '%Y-%m-%d') + timedelta(days=1)
                    query['updated_at'] = {
                        '$gte': start_date, '$lt': end_date}
                else:
                    single_date = datetime.strptime(
                        created_at_param, '%Y-%m-%d')
                    next_day = single_date + timedelta(days=1)
                    query['updated_at'] = {
                        '$gte': single_date, '$lt': next_day}
            except ValueError as e:
                return bad_request(f'Error al parsear la fecha: {created_at_param} — {e}')

        inbounds = InboundUseCase.get_by_external(self.material_id, query)
        if inbounds:
            return ok(self.__expand_items(inbounds))
        return ok([])

    def delete(self):
        with MongoDBHandler('inbounds') as db:
            material = db.extract(
                {'_id': ObjectId(self.id)}) if objectid_validation(self.id) else None
            if material:
                db.delete({'_id': ObjectId(self.id)})
                return ok('Entrada eliminada correctamente.')
            return bad_request('La entrada no existe.')

    def upload(self):
        required_fields = ['supplier_id', 'file']
        data = dict(self.data.items())
        if all(i in data for i in required_fields):
            serializer = FileUploadSerializer(data=self.data)
            if serializer.is_valid():
                excel_file = self.data['file']
                try:
                    workbook = load_workbook(excel_file, data_only=True)
                    items, errors = self.__process_workbook(
                        workbook, data['supplier_id'])
                    if len(items) > 0:
                        with MongoDBHandler('inbounds') as db:
                            data['folio'] = db.set_next_folio('inbound')
                            id = db.insert({
                                'purchase_order_id': self.__normalize_field(data['purchase_order_id']),
                                'supplier_id': data['supplier_id'],
                                'project': self.__normalize_field(data['project']),
                                'items': items,
                                'folio': data['folio'],
                                'notes': self.__normalize_field(data['notes']),
                                'status': 1,
                            })
                            self.__perform_counting(
                                db, self.__normalize_field(data['project']), items, id)
                    return created({
                        "message": f"Entradas procesadas correctamente: {len(items)} insertada(s) y hubó {len(errors)} error(es).",
                        "inserted": items,
                        "errors": errors,
                    })
                except Exception as e:
                    return bad_request(f'Error: {str(e)}, "Trace": {traceback.format_exc()}')
            return bad_request('Error al momento de cargar el arhivo Excel.')
        return bad_request('El proveedor así como el archivo en formato Excel son requeridos.')

    @staticmethod
    def check_quantities(items):
        for item in items:
            delivered_qty = item.get('delivered', {}).get('quantity', 0)
            total_qty = item.get('total_quantity', 0)

            if float(delivered_qty) < float(total_qty):
                return 1
        return 2

    @staticmethod
    def register(purchase_order_id, supplier_id, project, items, folio, notes):
        with MongoDBHandler('inbounds') as db:
            db.insert({
                'purchase_order_id': purchase_order_id,
                'supplier_id': supplier_id,
                'project': project,
                'items': items,
                'folio': folio,
                'notes': notes,
                'status': 0,
            })

    @staticmethod
    def get_by_external(material_id, query):
        with MongoDBHandler('inbounds') as db:
            inbounds = db.extract(query, 'updated_at', -1)
            if not inbounds:
                return []

            result = []
            for inbound in inbounds:
                filtered_items = [
                    item for item in inbound.get("items", [])
                    if item.get("material_id") == material_id
                ]
                if filtered_items:
                    inbound_copy = {**inbound, "items": filtered_items}
                    result.append(inbound_copy)

            return InboundSerializer(result, many=True).data
