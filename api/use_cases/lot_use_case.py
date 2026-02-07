import traceback
from bson import ObjectId
from collections import Counter
from openpyxl import load_workbook
from api_sataiga.handlers.mongodb_handler import MongoDBHandler
from api.helpers.http_responses import bad_request, ok, not_found
from api.serializers.lot_serializer import LotSerializer
from api.decorators.service_method import service_method
from api.helpers.validations import objectid_validation
from api.use_cases.explosion_use_case import ExplosionUseCase
from api.serializers.file_serializer import FileUploadSerializer
from api.services.lot_service import LotService


class LotUseCase:
    def __init__(self, request=None, **kwargs):
        if request:
            self.request = request
        self.data = kwargs.get('data', None)
        self.id = kwargs.get('id', None)
        self.home_production_id = kwargs.get('home_production_id', None)
        self.service = LotService()

    def __update_hp_lots(self, db, home_production_id):
        lots = db.extract({'home_production_id': home_production_id})
        if lots:
            prototype_counter = Counter()

            for lot in lots:
                prototype_counter[lot['prototype']] += 1

            result = {
                'total': len(lots),
                'prototypes': dict(prototype_counter),
            }

            total_sum = sum(float(l["percentage"]) for l in lots)
            MongoDBHandler.modify(db, 'home_production', {'_id': ObjectId(home_production_id)}, {
                'lots': result,
                'progress': round(total_sum / len(lots), 2)
            })

    def __update_explosion(self, home_production_id):
        explosion_use_case = ExplosionUseCase(
            home_production_id=home_production_id)
        explosion_use_case.create()

    def __get_prototypes(self):
        with MongoDBHandler('prototypes') as db:
            prototypes = db.extract(
                {'client_id': self.data['client_id'], 'front': self.data['front']})
            if prototypes:
                return [item["name"] for item in prototypes]
            return None

    def __process_workbook(self, workbook, existing_prototypes):
        valid_laid = ['IZQUIERDO', 'DERECHO']
        valid_prototypes = existing_prototypes

        valid_entries = []
        errors = []

        ws = workbook.active

        for row_index, row in enumerate(ws.iter_rows(min_row=2), start=2):
            entry = {}
            row_errors = []

            block = row[0].value
            lot = row[1].value
            laid = row[2].value
            prototype = row[3].value

            if block is None:
                row_errors.append("MANZANA vacía")
            else:
                entry["block"] = int(block)

            if lot is None:
                row_errors.append("LOTE vacío")
            else:
                entry["lot"] = int(lot)

            if laid is None or str(laid).strip().upper() not in valid_laid:
                row_errors.append(f"SEMBRADO inválido: {laid}")
            else:
                entry["laid"] = str(laid).strip().upper()

            if prototype is None or str(prototype).strip() not in valid_prototypes:
                row_errors.append(f"PROTOTIPO inválido: {prototype}")
            else:
                entry["prototype"] = str(prototype).strip()

            if row_errors:
                errors.append({"row": row_index, "errors": row_errors})
            else:
                valid_entries.append(entry)
        return valid_entries, errors

    @service_method()
    def save(self):
        return self.service.create(self.home_production_id, self.data)

    def get(self):
        with MongoDBHandler('lots') as db:
            lots = db.extract({'home_production_id': self.home_production_id})
            if lots:
                return ok(LotSerializer(lots, many=True).data)
            return ok([])

    def update(self):
        with MongoDBHandler('lots') as db:
            lot = db.extract(
                {'_id': ObjectId(self.id)}) if objectid_validation(self.id) else None
            if lot:
                db.update({'_id': ObjectId(self.id)}, self.data)
                self.__update_hp_lots(db, lot[0]['home_production_id'])
                return ok('Lote actualizado correctamente.')
            return not_found('El lote no existe.')

    def delete(self):
        with MongoDBHandler('lots') as db:
            lot = db.extract(
                {'_id': ObjectId(self.id)}) if objectid_validation(self.id) else None
            if lot:
                db.delete({'_id': ObjectId(self.id)})
                self.__update_hp_lots(db, lot[0]['home_production_id'])
                self.__update_explosion(lot[0]['home_production_id'])
                return ok('Lote eliminado correctamente.')
            return not_found('El lote no existe.')

    def upload(self):
        with MongoDBHandler('lots') as db:
            prototypes = self.__get_prototypes()
            if not prototypes:
                return bad_request('No existen prototipos para el cliente y el frente seleccionados.')

            serializer = FileUploadSerializer(data=self.data)
            if serializer.is_valid():
                excel_file = self.request.FILES['file']
                current_lots = db.extract(
                    {'home_production_id': self.home_production_id})
                if current_lots:
                    db.delete(
                        {'home_production_id': self.home_production_id})
                try:
                    workbook = load_workbook(excel_file, data_only=True)
                    lots, errors = self.__process_workbook(
                        workbook, prototypes)
                    for lot in lots:

                        db.insert({
                            'home_production_id': self.home_production_id,
                            **lot,
                            'percentage': 0,
                            'progress': {
                                'cocina': {
                                    'stages': [],
                                    'percentage': 0.0,
                                },
                                'closet': {
                                    'stages': [],
                                    'percentage': 0.0,
                                },
                                'puertas': {
                                    'stages': [],
                                    'percentage': 0.0,
                                },
                                'mdeb': {
                                    'stages': [],
                                    'percentage': 0.0,
                                },
                                'waldras': {
                                    'stages': [],
                                    'percentage': 0.0,
                                },
                                'instalacion': {
                                    'stages': [],
                                    'percentage': 0.0,
                                },
                            },
                        })
                    new_lots = db.extract(
                        {'home_production_id': self.home_production_id})
                    self.__update_hp_lots(db, self.home_production_id)
                    self.__update_explosion(self.home_production_id)
                    return ok({
                        'success': LotSerializer(new_lots, many=True).data,
                        'errors': errors,
                    })
                except Exception as e:
                    return bad_request(f'Error: {str(e)}, "Trace": {traceback.format_exc()}')
            return bad_request(serializer.errors)
