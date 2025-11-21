import os
import traceback
import qrcode
from api_sataiga.handlers.mongodb_handler import MongoDBHandler
from api.helpers.http_responses import created, bad_request, ok, ok_paginated, not_found
from bson import ObjectId
from api.helpers.validations import objectid_validation
from api.serializers.material_serializer import MaterialSerializer
from api.serializers.file_serializer import FileUploadSerializer
from openpyxl import load_workbook, Workbook
from decimal import Decimal, InvalidOperation
from rest_framework import exceptions
from django.http import HttpResponse
from PIL import Image
from django.conf import settings
from api.functions.concept_n_sku import generate_concept_and_sku
from api.services.material_service import MaterialService
from api.services.catalog_service import CatalogService
from api.utils.pagination_utils import DummyPaginator, DummyPage
from api.helpers.get_query_params import get_query_params


class MaterialUseCase:
    def __init__(self, request=None, **kwargs):
        params = get_query_params(request)
        self.q = params["q"]
        self.page = params["page"]
        self.page_size = params["page_size"]
        self.sort_by = params["sort_by"]
        self.order_by = params["order_by"]
        self.supplier_id = params.get('supplier_id', None)
        self.division = params.get('division', None)
        self.group = params.get('group', None)
        self.data = kwargs.get('data', None)
        self.id = kwargs.get('id', None)
        self.service = MaterialService()
        self.catalog_service = CatalogService()

    def __check_supplier(self, db):
        supplier = MongoDBHandler.find(db, 'suppliers', {'_id': ObjectId(
            self.data['supplier_id'])}) if objectid_validation(self.data['supplier_id']) else None
        if supplier:
            return True
        return False

    def __valid_value(self, data, idx):
        if len(data) <= idx:
            return None

        not_valid = ['', 'None']
        if not data[idx] or data[idx] in not_valid:
            return None

        if not isinstance(data[idx], str):
            return str(data[idx]).strip()

        return data[idx].strip()

    def __process_workbook(self, workbook):
        sheet = workbook.active

        data = []
        errors = []

        for idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            row_errors = []

            def get_decimal(value, field_name, idx=None, row_errors=None):
                try:
                    if value is None or str(value).strip() == '':
                        return None

                    cleaned = str(value).replace(
                        '$', '').replace(',', '').strip()

                    convert = Decimal(cleaned).quantize(Decimal('0.01'))
                    return str(convert)

                except (InvalidOperation, ValueError):
                    if row_errors is not None and idx is not None:
                        row_errors.append(
                            f"Fila {idx}: '{field_name}' no es un número decimal válido."
                        )
                    return None

            division = row[0]
            name = row[1]
            measurement = row[7]

            if not division or not name or not measurement:
                row_errors.append(
                    f"Fila {idx}: 'DIVISIÓN', 'MATERIAL' y 'UNIDAD DE MEDIDA' son obligatorios.")

            record = {
                "supplier_id": self.data['supplier_id'],
                "division": division.strip() if division else '',
                "name": name.strip() if name else '',
                "espec1": self.__valid_value(row, 2),
                "espec2": self.__valid_value(row, 3),
                "espec3": self.__valid_value(row, 4),
                "espec4": self.__valid_value(row, 5),
                "espec5": self.__valid_value(row, 6),
                "measurement": measurement.strip() if measurement else '',
                "supplier_code": self.__valid_value(row, 8),
                "presentation": self.__valid_value(row, 9),
                "area": self.__valid_value(row, 10),
                "reference": self.__valid_value(row, 11),
                "minimum": get_decimal(row[12], "MÍNIMOS DE STOCK"),
                "maximum": get_decimal(row[13], "MÁXIMOS DE STOCK"),
                "unit_price": get_decimal(row[14], "PRECIO UNIDAD"),
                "inventory_price": get_decimal(row[15], "PRECIO PRESENTACIÓN"),
                "market_price": get_decimal(row[16], "PRECIO MERCADO"),
                "price_difference": get_decimal(row[17], "DIFERENCIA DE PRECIO"),
                "automation": len(row) > 18 and str(row[18]) == 'SI'
            }

            if row_errors:
                errors.append({"fila": idx, "errores": row_errors})
            else:
                data.append(record)

        return data, errors

    def __process_materials(self, materials):
        with MongoDBHandler('materials') as db:
            inserted = []
            updated = []
            if self.__check_supplier(db):
                for item in materials:
                    filters = {
                        'supplier_id': item['supplier_id'],
                        'division': item['division'],
                        'name': item['name'],
                        'measurement': item['measurement'],
                    }
                    for f in ['espec1', 'espec2', 'espec3', 'espec4', 'espec5', 'supplier_code', 'presentation']:
                        if f in item and item[f]:
                            filters[f] = item[f]
                    material = db.extract(filters)
                    if material:
                        if 'sku' not in material[0] or not material[0]['sku']:
                            concept, sku = generate_concept_and_sku(
                                material[0])
                            item['concept'] = concept
                            item['sku'] = sku
                        if 'qr' not in material[0] or not material[0].get('qr'):
                            item['qr'] = self.__create_qr_image(
                                str(material[0]['_id']))
                        db.update({'_id': ObjectId(material[0]['_id'])}, item)
                        updated.append(
                            item['concept'] if 'concept' in item else material[0]['concept'])
                    else:
                        concept, sku = generate_concept_and_sku(item)
                        item['concept'] = concept
                        item['sku'] = sku
                        material_id = db.insert(item)
                        db.update(
                            {'_id': ObjectId(material_id)},
                            {'qr': self.__create_qr_image(str(material_id))})
                        inserted.append(item['concept'])
                return inserted, updated
            raise exceptions.NotFound('El proveedor seleccionado no existe.')

    def __suppliers_list(self):
        with MongoDBHandler('suppliers') as db:
            suppliers_list = {}
            suppliers = db.extract()
            if suppliers:
                for supplier in suppliers:
                    suppliers_list[str(supplier['_id'])] = supplier['name']
            return suppliers_list

    def __validate_value(self, material, key):
        if key in material:
            return material[key]
        return None

    def __export_materials(self, materials):
        suppliers_list = self.__suppliers_list()
        wb = Workbook()
        ws = wb.active
        ws.title = "Materiales"

        headers = [
            'PROVEEDOR', 'CONCEPTO', 'UNIDAD DE MEDIDA', 'CÓDIGO PROVEEDOR',
            'SKU', 'PRESENTACIÓN', 'ÁREA', 'REFERENCIA',
            'MÍNIMOS DE STOCK', 'MÁXIMOS DE STOCK',
            'PRECIO UNIDAD', 'PRECIO PRESENTACIÓN', 'PRECIO MERCADO', 'DIFERENCIA DE PRECIO'
        ]
        ws.append(headers)

        for material in materials:
            ws.append([
                suppliers_list[material['supplier_id']],
                material['concept'],
                material['measurement'],
                self.__validate_value(material, 'supplier_code'),
                self.__validate_value(material, 'sku'),
                self.__validate_value(material, 'presentation'),
                self.__validate_value(material, 'area'),
                self.__validate_value(material, 'reference'),
                float(self.__validate_value(material, 'minimum') or 0),
                float(self.__validate_value(material, 'maximum') or 0),
                float(self.__validate_value(material, 'unit_price') or 0),
                float(self.__validate_value(material, 'invetory_price') or 0),
                float(self.__validate_value(material, 'market_price') or 0),
                float(self.__validate_value(material, 'price_difference') or 0),
            ])

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename=materiales.xlsx'
        wb.save(response)
        return response

    def __create_qr_image(self, material_id):
        # Construir la URL
        url = f"{settings.ADMIN_URL}apps/materials/view/{material_id}?input=true"

        # Ruta donde se guardará la imagen
        qr_dir = os.path.join(settings.MEDIA_ROOT, "materials/qr")
        os.makedirs(qr_dir, exist_ok=True)  # Crea la carpeta si no existe

        qr_path = os.path.join(qr_dir, f"{material_id}.jpg")

        # Crear el QR code
        qr = qrcode.QRCode(
            version=1,  # controla el tamaño de la matriz
            error_correction=qrcode.constants.ERROR_CORRECT_H,  # alta tolerancia de error
            box_size=10,  # tamaño de cada "cuadro"
            border=4,  # borde alrededor
        )
        qr.add_data(url)
        qr.make(fit=True)

        # Generar imagen
        img = qr.make_image(fill_color="black", back_color="white")
        img.save(qr_path)

        return f"{settings.BASE_URL}/media/materials/qr/{material_id}.jpg"

    def __build_material_filters(self, division_list: list[str] = None):
        """
        Construye el filtro MongoDB para materiales, considerando búsqueda,
        proveedor, división, y grupo (EQUIPMENT_GROUP / MATERIALS_GROUP).
        """
        filters = {}

        # 🔹 Búsqueda libre
        if self.q:
            filters['$or'] = [
                {'concept': {'$regex': self.q, '$options': 'i'}},
                {'measurement': {'$regex': self.q, '$options': 'i'}},
                {'supplier_code': {'$regex': self.q, '$options': 'i'}},
                {'sku': {'$regex': self.q, '$options': 'i'}},
                {'presentation': {'$regex': self.q, '$options': 'i'}},
                {'reference': {'$regex': self.q, '$options': 'i'}},
            ]

        # 🔹 Filtro de proveedor (prioriza el override si se pasa explícito)
        if self.supplier_id:
            filters['supplier_id'] = self.supplier_id

        # 🔹 Si viene una lista de divisiones explícita
        if division_list:
            filters['division'] = {'$in': division_list}
            return filters  # ← No aplicar filtros por grupo

        # 🔹 Filtro directo de división (tiene prioridad)
        if self.division:
            filters['division'] = self.division
            return filters  # ← También evita aplicar group

        # 🔹 Filtro por grupo (solo si no hay división)
        if getattr(self, "group", None):
            group_name = self.group.upper()

            # Consultar catálogo dinámico desde MongoDB
            catalog = self.catalog_service.get_by_name(
                'Equipos y/o accesorios')
            equipment_divisions = catalog.get('values', []) if catalog else []

            if group_name == "EQUIPMENT_GROUP":
                filters['division'] = {'$in': equipment_divisions}
            elif group_name == "MATERIALS_GROUP":
                filters['division'] = {'$nin': equipment_divisions}

        return filters

    def save(self):
        with MongoDBHandler('materials') as db:
            required_fields = ['division', 'name',
                               'concept', 'measurement', 'supplier_id', 'sku']
            if all(i in self.data for i in required_fields):
                if self.__check_supplier(db):
                    if 'automation' not in self.data:
                        self.data['automation'] = False
                    material_id = db.insert(self.data)
                    db.update(
                        {'_id': ObjectId(material_id)},
                        {'qr': self.__create_qr_image(str(material_id))})
                    return created({'id': str(material_id)})
                return bad_request('El proveedor selecionado no existe.')
            return bad_request('Algunos campos requeridos no han sido completados.')

    def get(self):
        try:
            division_list = None
            if self.division and len(self.division) > 0:
                division_list = self.division.split(',')

            filters = self.__build_material_filters(division_list)
            result = self.service.get_paginated(
                filters, self.page, self.page_size, self.sort_by, self.order_by
            )

            dummy_paginator = DummyPaginator(
                result["count"], result["total_pages"])
            dummy_page = DummyPage(
                result["current_page"], dummy_paginator, result["results"]
            )

            return ok_paginated(dummy_paginator, dummy_page, result["results"])

        except Exception as e:
            return bad_request(f"Error al obtener prototipos: {e}")

    def get_by_id(self):
        with MongoDBHandler('materials') as db:
            material = db.extract(
                {'_id': ObjectId(self.id)}) if objectid_validation(self.id) else None
            if material:
                return ok(MaterialSerializer(material[0]).data)
            return not_found('El material no existe.')

    def update(self):
        with MongoDBHandler('materials') as db:
            material = db.extract(
                {'_id': ObjectId(self.id)}) if objectid_validation(self.id) else None
            if len(material) > 0:
                if 'supplier_id' in self.data and not self.__check_supplier(db):
                    return bad_request('El proveedor selecionado no existe.')
                if 'qr' not in material[0] or not material[0].get('qr'):
                    self.data['qr'] = self.__create_qr_image(
                        str(material[0]['_id']))
                updated = db.update({'_id': ObjectId(self.id)}, self.data)
                return ok(MaterialSerializer(updated[0]).data)
            return bad_request('El material no existe.')

    def delete(self):
        with MongoDBHandler('materials') as db:
            material = db.extract(
                {'_id': ObjectId(self.id)}) if objectid_validation(self.id) else None
            if material:
                db.delete({'_id': ObjectId(self.id)})
                return ok('Material eliminado correctamente.')
            return bad_request('El material no existe.')

    def upload(self):
        required_fields = ['supplier_id', 'file']
        data = dict(self.data.items())
        if all(i in data for i in required_fields):
            serializer = FileUploadSerializer(data=self.data)
            if serializer.is_valid():
                excel_file = self.data['file']
                try:
                    workbook = load_workbook(excel_file, data_only=True)
                    materials, errors = self.__process_workbook(workbook)
                    inserted, updated = self.__process_materials(materials)
                    return ok({
                        "message": f"Materiales procesados correctamente: {len(inserted)} insertado(s), {len(updated)} atualizado(s) y hubó {len(errors)} error(es).",
                        "inserted": inserted,
                        "updated": updated,
                        "errors": errors,
                    })
                except Exception as e:
                    return bad_request(f'Error: {str(e)}, "Trace": {traceback.format_exc()}')
            return bad_request('Error al momento de cargar el arhivo Excel.')
        return bad_request('El proveedor así como el archivo en formato Excel son requeridos.')

    def download(self):
        with MongoDBHandler('materials') as db:
            # Convertir divisiones separadas por coma a lista
            division_list = None
            if self.division and len(self.division) > 0:
                division_list = self.division.split(',')

            # Reutilizamos la misma lógica de construcción de filtros
            filter_query = self.__build_material_filters(
                division_list=division_list)
            materials = db.extract(filter_query)
            return self.__export_materials(materials)

    def upload_image(self):
        with MongoDBHandler('materials') as db:
            images = []
            image = self.request.FILES['image']
            material = db.extract(
                {'_id': ObjectId(self.id)}) if objectid_validation(self.id) else None
            if material and image:
                if 'images' in material[0]:
                    images = material[0]['images']

                img = Image.open(image)
                if img.format not in ['JPEG', 'PNG']:
                    return bad_request('El archivo cargado no es una imagen valida.')

                material_folder = os.path.join(
                    settings.MEDIA_ROOT, 'materials/' + str(self.id))
                os.makedirs(material_folder, exist_ok=True)

                image_path = os.path.join(material_folder, image.name)

                with open(image_path, 'wb+') as destination:
                    for chunk in image.chunks():
                        destination.write(chunk)

                relative_path = f"{settings.BASE_URL}/media/materials/{str(self.id)}/{image.name}"
                images.append(relative_path)

                db.update({'_id': ObjectId(self.id)}, {'images': images})

                return ok(images)
            return bad_request('El material no existe.')

    def delete_image(self):
        with MongoDBHandler('materials') as db:
            if 'images' in self.data:
                material = db.extract(
                    {'_id': ObjectId(self.id)}) if objectid_validation(self.id) else None
                if material:
                    if 'images' in material[0] and isinstance(self.data['images'], list):
                        deleted = []
                        for idx in self.data['images']:
                            deleted.append(material[0]['images'][idx])
                            relative_path = material[0]['images'][idx].replace(
                                settings.BASE_URL, '')
                            file_path = os.path.join(
                                settings.BASE_DIR, relative_path.strip('/'))
                            if os.path.isfile(file_path):
                                os.remove(file_path)

                        new_images = [url for url in material[0]
                                      ['images'] if url not in deleted]
                        db.update({'_id': ObjectId(self.id)}, {
                            'images': new_images})
                        return ok(new_images)
                    return bad_request('No existen imágenes en el material seleccionado y/o las imágenes seleccionadas para eliminar.')
                return bad_request('El material no existe.')
            return bad_request('Falta el índice de la imagen para poder eliminarla.')
