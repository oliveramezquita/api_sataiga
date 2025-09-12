import traceback
from api_sataiga.handlers.mongodb_handler import MongoDBHandler
from api.helpers.validations import objectid_validation
from bson import ObjectId
from api.helpers.http_responses import bad_request, ok, not_found
from api.serializers.volumetry_serializer import VolumetrySerializer
from api.serializers.file_serializer import FileUploadSerializer
from urllib.parse import parse_qs
from openpyxl import load_workbook
from api.functions.quantify import quantify


class VolumetryUseCase:
    def __init__(self, request=None, **kwargs):
        if request:
            self.request = request
            params = parse_qs(request.META['QUERY_STRING'])
            self.client_id = params['client_id'][0] if 'client_id' in params else None
            self.front = params['front'][0] if 'front' in params else None
            self.prototype = params['prototype'][0] if 'prototype' in params else None
        self.data = kwargs.get('data', None)
        self.id = kwargs.get('id', None)

    def __client_validation(self, db, client_id):
        client = MongoDBHandler.find(db, 'clients', {'_id': ObjectId(
            client_id), 'type': 'VS'}) if objectid_validation(client_id) else None
        if client:
            return client[0]
        return False

    def __material_validation(self, db):
        material = MongoDBHandler.find(db, 'materials', {'_id': ObjectId(
            self.data['material_id'])}) if objectid_validation(self.data['material_id']) else None
        if material:
            return material
        return False

    def __calculate_total(self, reference):
        return {
            'volumetry': [item.update({'total_x': float(item['total_x'])}) or item for item in self.data['volumetry']],
            'total': sum(float(item['total_x']) for item in self.data['volumetry']),
            'reference': reference,
        }

    def __get_id_material_by_sku(self, sku):
        with MongoDBHandler('materials') as db:
            if not sku:
                return False
            material = db.extract({'sku': sku})
            if material:
                return str(material[0]['_id'])
            return False

    def __process_workbook_general(self, wb):
        ws = wb.active

        # Detectar áreas desde la fila 1 y 2
        areas = []
        col = 5  # Columna E
        while ws.cell(row=1, column=col).value:
            area_name = ws.cell(row=1, column=col).value
            factory_header = ws.cell(row=2, column=col).value
            install_header = ws.cell(row=2, column=col + 1).value

            areas.append({
                "name": area_name,
                "factory_col": col,
                "install_col": col + 1,
                "factory_header": factory_header,
                "install_header": install_header,
            })

            col += 2

        results = []
        errors = []

        # Procesar filas a partir de la fila 3
        for row in range(3, ws.max_row + 1):
            material = ws.cell(row=row, column=1).value
            sku = ws.cell(row=row, column=3).value
            reference = ws.cell(row=row, column=4).value

            # Validar campos obligatorios
            if not sku:
                errors.append({
                    "row": row,
                    "message": f"Error en el material: {material or 'DESCONOCIDO'}, faltan valores obligatorios (C)."
                })
                continue

            material_id = self.__get_id_material_by_sku(sku)
            if not material_id:
                errors.append({
                    "row": row,
                    "message": f"El material: {material} con el SKU: {sku} no existe en la base de datos, verifique que el SKU sea correcto."
                })
                continue

            volumetry_list = []

            for area in areas:
                factory_val = ws.cell(
                    row=row, column=area["factory_col"]).value
                install_val = ws.cell(
                    row=row, column=area["install_col"]).value

                def parse_number(val):
                    if val is None or val == "":
                        return None
                    try:
                        return round(float(val), 2)
                    except (ValueError, TypeError):
                        return "INVALID"

                factory = parse_number(factory_val)
                install = parse_number(install_val)

                # Si alguno es inválido, registrar error
                if factory == "INVALID" or install == "INVALID":
                    errors.append({
                        "row": row,
                        "message": f"Error en el material: {material}, revisar los datos ingresados en área {area['name']}."
                    })
                    continue

                if factory is not None or install is not None:
                    total_x = (factory or 0) + (install or 0)
                    volumetry_list.append({
                        "area": area["name"],
                        "factory": factory,
                        "installation": install,
                        "total_x": round(total_x, 2)
                    })

            # Validar que al menos una área tenga valores
            if not volumetry_list:
                errors.append({
                    "row": row,
                    "message": f"Error en el material: {material}, no se encontró ninguna área con cantidades."
                })
                continue

            results.append({
                'material_id': material_id,
                'reference': reference,
                'volumetry': volumetry_list,
                'total': round(sum(float(vl['total_x']) for vl in volumetry_list), 2)
            })

        return results, errors

    def __process_workbook_kitchen(self, wb):
        ws = wb.active

        # Detectar áreas desde la fila 1 y 2 (ahora en bloques de 3 columnas)
        areas = []
        col = 5  # Columna E
        while ws.cell(row=1, column=col).value:
            area_name = ws.cell(row=1, column=col).value
            factory_header = ws.cell(row=2, column=col).value
            install_header = ws.cell(row=2, column=col + 1).value
            delivery_header = ws.cell(row=2, column=col + 2).value

            areas.append({
                "name": area_name,
                "factory_col": col,
                "install_col": col + 1,
                "delivery_col": col + 2,
                "factory_header": factory_header,
                "install_header": install_header,
                "delivery_header": delivery_header,
            })

            col += 3

        results = []
        errors = []

        # Procesar filas a partir de la fila 3
        for row in range(3, ws.max_row + 1):
            material = ws.cell(row=row, column=1).value
            sku = ws.cell(row=row, column=3).value
            reference = ws.cell(row=row, column=4).value

            # Validar campos obligatorios
            if not sku:
                errors.append({
                    "row": row,
                    "message": f"Error en el material: {material or 'DESCONOCIDO'}, faltan valores obligatorios (C)."
                })
                continue

            material_id = self.__get_id_material_by_sku(sku)
            if not material_id:
                errors.append({
                    "row": row,
                    "message": f"El material: {material} con el SKU: {sku} no existe en la base de datos, verifique que el SKU sea correcto."
                })
                continue

            volumetry_list = []

            for area in areas:
                factory_val = ws.cell(
                    row=row, column=area["factory_col"]).value
                install_val = ws.cell(
                    row=row, column=area["install_col"]).value
                delivery_val = ws.cell(
                    row=row, column=area["delivery_col"]).value

                def parse_number(val):
                    if val is None or val == "":
                        return None
                    try:
                        return round(float(val), 2)
                    except (ValueError, TypeError):
                        return "INVALID"

                factory = parse_number(factory_val)
                install = parse_number(install_val)
                delivery = parse_number(delivery_val)

                # Si alguno es inválido, registrar error
                if factory == "INVALID" or install == "INVALID" or delivery == "INVALID":
                    errors.append({
                        "row": row,
                        "message": f"Error en el material: {material}, revisar los datos ingresados en área {area['name']}."
                    })
                    continue

                if factory is not None or install is not None or delivery is not None:
                    total_x = (factory or 0) + (install or 0) + (delivery or 0)
                    volumetry_list.append({
                        "area": area["name"],
                        "factory": factory,
                        "installation": install,
                        "delivery": delivery,
                        "total_x": round(total_x, 2)
                    })

            # Validar que al menos una área tenga valores
            if not volumetry_list:
                errors.append({
                    "row": row,
                    "message": f"Error en el material: {material}, no se encontró ninguna área con cantidades."
                })
                continue

            results.append({
                'material_id': material_id,
                'reference': reference,
                'volumetry': volumetry_list,
                'total': round(sum(float(vl['total_x']) for vl in volumetry_list), 2)
            })

        return results, errors

    def __process_data(self, volumetry_data):
        updated = 0
        inserted = 0
        with MongoDBHandler('volumetries') as db:
            for item in volumetry_data:
                is_exist = db.extract(
                    {'client_id': self.client_id, 'front': self.front, 'prototype': self.prototype, 'material_id': item['material_id']})
                if self.__client_validation(db, self.client_id):
                    if is_exist:
                        db.update({
                            'client_id': self.client_id,
                            'front': self.front,
                            'prototype': self.prototype,
                            'material_id': item['material_id']},
                            {
                                'volumetry': item['volumetry'],
                                'total': item['total'],
                                'reference': item['reference']
                        })
                        updated += 1
                    else:
                        db.insert({
                            'client_id': self.client_id,
                            'front': self.front,
                            'prototype': self.prototype,
                            **item
                        })
                        inserted += 1
        return inserted, updated

    def __extract(self):
        with MongoDBHandler('volumetries') as db:
            return db.extract(
                {'client_id': self.client_id, 'front': self.front, 'prototype': self.prototype})

    def save(self):
        with MongoDBHandler('volumetries') as db:
            required_fields = ['client_id', 'front',
                               'prototype', 'material_id', 'volumetry']
            reference = self.data['reference'] if 'reference' in self.data else None
            if all(i in self.data for i in required_fields):
                material = self.__material_validation(db)
                if self.__client_validation(db, self.data['client_id']) and material:
                    is_exist = db.extract(
                        {
                            'client_id': self.data['client_id'],
                            'front': self.data['front'],
                            'prototype': self.data['prototype'],
                            'material_id': self.data['material_id']
                        })
                    if is_exist:
                        db.update({
                            'client_id': self.data['client_id'],
                            'front': self.data['front'],
                            'prototype': self.data['prototype'],
                            'material_id': self.data['material_id']},
                            self.__calculate_total(reference)
                        )
                        message = f'El material: {material[0]['concept']} ha sido actualizado correctamente en la volumetría.'
                    else:
                        db.insert({
                            'client_id': self.data['client_id'],
                            'front': self.data['front'],
                            'prototype': self.data['prototype'],
                            'material_id': self.data['material_id'],
                            **self.__calculate_total(reference),
                        })
                        message = f'El material: {material[0]['concept']} ha sido añadido correctamente en la volumetría.'
                    volumetry = db.extract(
                        {
                            'client_id': self.data['client_id'],
                            'front': self.data['front'],
                            'prototype': self.data['prototype'],
                        })
                    quantify.delay(
                        self.data['client_id'],
                        self.data['front'],
                        self.data['prototype'],
                        VolumetrySerializer(volumetry, many=True).data)
                    return ok({'data': VolumetrySerializer(volumetry, many=True).data, 'message': message})
                return bad_request('Error al momento de procesar la información: el cliente o el material no existen.')
            return bad_request('Algunos campos requeridos no han sido completados.')

    def get(self):
        if self.client_id and self.front and self.prototype:
            with MongoDBHandler('volumetries') as db:
                volumetry = db.extract({
                    'client_id': self.client_id,
                    'front': self.front,
                    'prototype': self.prototype,
                })
                return ok({
                    'volumetry': VolumetrySerializer(volumetry, many=True).data,
                })
        return not_found('No existe volumería con lo datos asignados.')

    def delete(self):
        with MongoDBHandler('volumetries') as db:
            volumetry = db.extract(
                {'_id': ObjectId(self.id)}) if objectid_validation(self.id) else None
            if volumetry:
                db.delete({'_id': ObjectId(self.id)})
                return ok('El área en la volumetría ha sido eliminado correctamente.')
            return bad_request('El área en la volumetría no existe.')

    def upload(self):
        serializer = FileUploadSerializer(data=self.data)
        if not serializer.is_valid():
            return bad_request(serializer.errors)

        excel_file = self.request.FILES['file']

        try:
            workbook = load_workbook(excel_file, data_only=True)

            # Selección de función según el prototipo
            if self.prototype.strip().endswith("Cocina"):
                volumetry_data, errors = self.__process_workbook_kitchen(
                    workbook)
            else:
                volumetry_data, errors = self.__process_workbook_general(
                    workbook)

            # Procesar datos
            num_inserted, num_updated = self.__process_data(volumetry_data)

            volumetry = []
            if num_inserted > 0 or num_updated > 0:
                volumetry = VolumetrySerializer(
                    self.__extract(), many=True
                ).data
                # Llamada asíncrona a Celery
                quantify.delay(
                    self.client_id,
                    self.front,
                    self.prototype,
                    volumetry
                )

            return ok({
                'num_inserted': num_inserted,
                'num_updated': num_updated,
                'errors': errors,
                'volumetry': volumetry,
            })

        except Exception as e:
            return bad_request({
                'error': str(e),
                'trace': traceback.format_exc()
            })
